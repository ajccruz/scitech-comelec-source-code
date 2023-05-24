#!/usr/bin/env python3
import sqlite3

import openpyxl

def getElectionResults(dbFile):
    con = sqlite3.connect(dbFile)
    cur = con.cursor()
    results = cur.execute("SELECT * FROM Candidates").fetchall()
    out = {}
    #The following tuple indices hold the respective fields:   0: name, 1: office, 2: party, 3: votes
    for row in results:
        office = row[1]
        if office not in out:
            out[office] = {}
        
        out[office][row[0]] = row[3]
        
    return out

def saveResultsToExcel(results, output="winners.xlsx"):
    wb = openpyxl.Workbook()
    sheet = wb.active

    row = 1
    for office in results:
        sheet.cell(row=row, column=1).value = office
        for candidate, votes in results[office].items():
            column = 2
            sheet.cell(row=row, column=column).value = candidate
            column += 1
            sheet.cell(row=row, column=column).value = votes
            row += 1

        row += 2

    wb.save(output)

if __name__ == "__main__":
    for office, results in getElectionResults("../votesDB.db").items():
        print(office)
        for candidate, votes in results.items():
            print(candidate, ": ", votes, sep="")

        print()

        saveResultsToExcel(getElectionResults("../votesDB.db"), "../winners.xlsx")